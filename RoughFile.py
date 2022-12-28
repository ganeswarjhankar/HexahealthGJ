##Necessary webdrivers ned to be imported
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
#from selenium.webdriver.common.by import By

from openpyxl import load_workbook

S = Service("D:\\chromedriver.exe")
driver = webdriver.Chrome(service=S)

# Load the workbook and select the sheet
wb = load_workbook("C:\\Users\\91928\\PycharmProjects\\MultiUrlsFile.xlsx")
sheet = wb["Sheet1"]

# Iterate over the cells in the sheet
for i in range(2, sheet.max_row + 1):
    # Read the URL from the cell
    url = sheet.cell(row=i, column=1).value

    # Use the URL in a Selenium script
    #driver = webdriver.Chrome()
    driver.get(url)
    # Perform other actions with Selenium here
    #driver.quit()

# Close the webdriver
#driver.quit()




