"""Open all the Urls in one go with different tabs"""

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
S=Service("D:\\chromedriver.exe")
driver=webdriver.Chrome(service=S)

# Open the Excel sheet
wb = load_workbook('C:\\Users\\91928\\PycharmProjects\\MultiUrlsFile.xlsx')

# Select the sheet containing the URLs
sheet = wb['Sheet1']

# Read the URLs into a list
urls = []
for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
    url = row[0]
    urls.append(url)



#########################################################

# Iterate through the list of URLs
for url in urls:
    # Open the URL in a new window
    driver.execute_script("window.open('{}');".format(url))




# Close the webdriver when you're done
driver.close()
